"""Generate the change request Excel template.

Run once to produce change_request_template.xlsx in this directory.
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

wb = Workbook()
ws = wb.active
ws.title = "Change Request"

# ---------------------------------------------------------------------------
# Styles
# ---------------------------------------------------------------------------
NAVY = "1F3864"
BLUE = "4472C4"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
LIGHT_GRAY = "F2F2F2"
GREEN = "C6EFCE"
YELLOW = "FFEB9C"
RED = "FFC7CE"

header_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
header_fill = PatternFill(start_color=BLUE, end_color=BLUE, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

title_font = Font(name="Calibri", bold=True, color=WHITE, size=14)
title_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")

meta_label_font = Font(name="Calibri", bold=True, size=10)
meta_value_font = Font(name="Calibri", size=10)

thin_border = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)

# ---------------------------------------------------------------------------
# Title banner (rows 1-2, merged)
# ---------------------------------------------------------------------------
ws.merge_cells("A1:J2")
title_cell = ws["A1"]
title_cell.value = "FIREWALL CHANGE REQUEST"
title_cell.font = title_font
title_cell.fill = title_fill
title_cell.alignment = Alignment(horizontal="center", vertical="center")

# ---------------------------------------------------------------------------
# Metadata section (rows 3-7)
# ---------------------------------------------------------------------------
meta_fields = [
    ("A3", "Change Request #:", "B3", ""),
    ("A4", "Requested By:", "B4", ""),
    ("A5", "Date:", "B5", ""),
    ("D3", "Approved By:", "E3", ""),
    ("D4", "Approval Date:", "E4", ""),
    ("D5", "Status:", "E5", ""),
    ("G3", "Environment:", "H3", ""),
    ("G4", "Priority:", "H4", ""),
    ("G5", "Ticket/Ref:", "H5", ""),
]

for label_ref, label_text, value_ref, value_text in meta_fields:
    label_cell = ws[label_ref]
    label_cell.value = label_text
    label_cell.font = meta_label_font
    label_cell.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

    value_cell = ws[value_ref]
    value_cell.value = value_text
    value_cell.font = meta_value_font
    value_cell.border = Border(bottom=Side(style="thin", color=BLUE))

# Description row
ws.merge_cells("A6:J6")
desc_label = ws["A6"]
desc_label.value = "Description:"
desc_label.font = meta_label_font
desc_label.fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")

ws.merge_cells("A7:J7")
desc_cell = ws["A7"]
desc_cell.value = ""
desc_cell.font = meta_value_font
desc_cell.border = Border(bottom=Side(style="thin", color=BLUE))
desc_cell.alignment = Alignment(wrap_text=True)
ws.row_dimensions[7].height = 40

# Row 8 is blank separator
ws.row_dimensions[8].height = 8

# ---------------------------------------------------------------------------
# Rule table — headers on row 9 (but we also put them on a hidden row 1 of
# a second sheet so the fill_template code can find them on row 1)
# ---------------------------------------------------------------------------
columns = [
    ("Source", 22),
    ("Destination", 22),
    ("Port", 12),
    ("Protocol", 12),
    ("Direction", 14),
    ("Action", 12),
    ("Justification", 45),
    ("Risk", 12),
    ("Proposal ID", 16),
    ("Approval Required", 20),
]

HEADER_ROW = 9
for col_idx, (name, width) in enumerate(columns, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col_idx, value=name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws.column_dimensions[get_column_letter(col_idx)].width = width

# Pre-format 50 data rows with alternating fills and borders
DATA_START = HEADER_ROW + 1
for row_offset in range(50):
    row_num = DATA_START + row_offset
    fill_color = LIGHT_GRAY if row_offset % 2 == 0 else WHITE
    row_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.fill = row_fill
        cell.border = thin_border
        cell.alignment = Alignment(vertical="top", wrap_text=True)

# ---------------------------------------------------------------------------
# Data validation dropdowns
# ---------------------------------------------------------------------------
action_dv = DataValidation(
    type="list",
    formula1='"ALLOW,DENY,MODIFY,RESTRICT"',
    allow_blank=True,
)
action_dv.error = "Please select a valid action"
action_dv.errorTitle = "Invalid Action"
ws.add_data_validation(action_dv)
action_dv.add(f"F{DATA_START}:F{DATA_START + 49}")

risk_dv = DataValidation(
    type="list",
    formula1='"LOW,MEDIUM,HIGH,CRITICAL"',
    allow_blank=True,
)
risk_dv.error = "Please select a valid risk level"
risk_dv.errorTitle = "Invalid Risk"
ws.add_data_validation(risk_dv)
risk_dv.add(f"H{DATA_START}:H{DATA_START + 49}")

direction_dv = DataValidation(
    type="list",
    formula1='"INBOUND,OUTBOUND"',
    allow_blank=True,
)
direction_dv.error = "Please select a valid direction"
direction_dv.errorTitle = "Invalid Direction"
ws.add_data_validation(direction_dv)
direction_dv.add(f"E{DATA_START}:E{DATA_START + 49}")

protocol_dv = DataValidation(
    type="list",
    formula1='"TCP,UDP,ICMP,ALL"',
    allow_blank=True,
)
protocol_dv.error = "Please select a valid protocol"
protocol_dv.errorTitle = "Invalid Protocol"
ws.add_data_validation(protocol_dv)
protocol_dv.add(f"D{DATA_START}:D{DATA_START + 49}")

approval_dv = DataValidation(
    type="list",
    formula1='"Yes,No"',
    allow_blank=True,
)
ws.add_data_validation(approval_dv)
approval_dv.add(f"J{DATA_START}:J{DATA_START + 49}")

# ---------------------------------------------------------------------------
# Freeze panes below header row
# ---------------------------------------------------------------------------
ws.freeze_panes = f"A{DATA_START}"

# ---------------------------------------------------------------------------
# Data sheet — hidden sheet with row-1 headers for fill_template compatibility
# ---------------------------------------------------------------------------
data_ws = wb.create_sheet("_data")
for col_idx, (name, _) in enumerate(columns, start=1):
    data_ws.cell(row=1, column=col_idx, value=name)
data_ws.sheet_state = "hidden"

# ---------------------------------------------------------------------------
# Print setup
# ---------------------------------------------------------------------------
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = "landscape"
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0

# ---------------------------------------------------------------------------
# Save
# ---------------------------------------------------------------------------
output = "change_request_template.xlsx"
wb.save(output)
print(f"Template saved: {output}")
