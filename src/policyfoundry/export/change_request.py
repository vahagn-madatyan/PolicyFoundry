"""Change request form export (xlsx + pdf).

Generates styled xlsx workbooks and PDF documents from ChangeRequestEntry
data, supporting both default template generation and user-provided
template fill for xlsx.
"""

from __future__ import annotations

from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fpdf import FPDF
from fpdf.fonts import FontFace
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from policyfoundry.exceptions import ExportError
from policyfoundry.export.models import ChangeRequestEntry
from policyfoundry.pipeline.excel_state import ExcelPipelineState
from policyfoundry.export.models import flatten_to_entries


# Column name → ChangeRequestEntry field mapping (case-insensitive keys)
COLUMN_MAP: dict[str, str] = {
    "source": "source",
    "destination": "destination",
    "port": "port",
    "protocol": "protocol",
    "direction": "direction",
    "action": "action",
    "justification": "justification",
    "risk": "risk",
    "proposal id": "proposal_id",
    "proposal_id": "proposal_id",
    "approval required": "approval_required",
    "approval_required": "approval_required",
}

# Default column order and display headers
_DEFAULT_COLUMNS: list[tuple[str, str, int]] = [
    ("source", "Source", 22),
    ("destination", "Destination", 22),
    ("port", "Port", 10),
    ("protocol", "Protocol", 10),
    ("direction", "Direction", 12),
    ("action", "Action", 10),
    ("justification", "Justification", 40),
    ("risk", "Risk", 10),
    ("proposal_id", "Proposal ID", 15),
    ("approval_required", "Approval Required", 18),
]


def _write_metadata(
    ws: Worksheet,
    state: ExcelPipelineState,
    entry_count: int,
) -> int:
    """Write metadata header rows (rows 1–5) and return first data row.

    Returns:
        Row number of the first available data row (after blank separator).
    """
    ws.cell(row=1, column=1, value="Generated")
    ws.cell(row=1, column=2, value=state.get("started_at", ""))

    ws.cell(row=2, column=1, value="Run ID")
    ws.cell(row=2, column=2, value=state.get("run_id", ""))

    ws.cell(row=3, column=1, value="Source Type")
    ws.cell(row=3, column=2, value="Excel Pipeline")

    ws.cell(row=4, column=1, value="Total Rules")
    ws.cell(row=4, column=2, value=entry_count)

    # Row 5 is a blank separator
    # Bold the metadata labels
    bold_font = Font(bold=True)
    for row in range(1, 5):
        cell = ws.cell(row=row, column=1)
        cell.font = bold_font

    return 6  # header row starts at 6


def _write_default_sheet(
    ws: Worksheet,
    entries: list[ChangeRequestEntry],
    header_row: int,
) -> None:
    """Write styled header and data rows to the worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)

    # Write header row
    for col_idx, (field, display_name, width) in enumerate(_DEFAULT_COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=display_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        # Set column width using letter
        col_letter = chr(64 + col_idx) if col_idx <= 26 else None
        if col_letter:
            ws.column_dimensions[col_letter].width = width

    # Write data rows
    for row_offset, entry in enumerate(entries):
        row_num = header_row + 1 + row_offset
        entry_dict = entry.model_dump()
        for col_idx, (field, _, _) in enumerate(_DEFAULT_COLUMNS, start=1):
            value = entry_dict[field]
            # Convert bool to Yes/No for display
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            ws.cell(row=row_num, column=col_idx, value=value)


def _fill_template(
    ws: Worksheet,
    entries: list[ChangeRequestEntry],
) -> None:
    """Fill a user-provided template by matching column headers.

    Scans row 1 for known column names (case-insensitive), finds the first
    empty row after the header, and writes data into matched columns only.
    """
    # Scan header row for column mapping
    col_mapping: dict[int, str] = {}  # col_index → field_name
    for col_idx in range(1, ws.max_column + 1 if ws.max_column else 1):
        header_val = ws.cell(row=1, column=col_idx).value
        if header_val is not None:
            normalized = str(header_val).strip().lower()
            if normalized in COLUMN_MAP:
                col_mapping[col_idx] = COLUMN_MAP[normalized]

    if not col_mapping:
        return  # No matching columns found — nothing to fill

    # Find first empty row after header
    start_row = 2
    for row_num in range(2, (ws.max_row or 1) + 1):
        has_data = any(
            ws.cell(row=row_num, column=c).value is not None
            for c in col_mapping
        )
        if not has_data:
            start_row = row_num
            break
    else:
        start_row = (ws.max_row or 1) + 1

    # Write data
    for row_offset, entry in enumerate(entries):
        row_num = start_row + row_offset
        entry_dict = entry.model_dump()
        for col_idx, field_name in col_mapping.items():
            value = entry_dict[field_name]
            if isinstance(value, bool):
                value = "Yes" if value else "No"
            ws.cell(row=row_num, column=col_idx, value=value)


def export_xlsx(
    state: ExcelPipelineState,
    output_path: str | Path,
    *,
    template_path: str | Path | None = None,
) -> Path:
    """Export pipeline state as an xlsx change request form.

    Args:
        state: ExcelPipelineState dict from pipeline execution.
        output_path: Destination path for the generated xlsx file.
        template_path: Optional path to a user-provided xlsx template.
            When provided, data is filled into matching columns.
            When None, a default styled workbook is generated.

    Returns:
        Path to the written xlsx file.

    Raises:
        ExportError: If workbook creation or writing fails.
    """
    output = Path(output_path)
    entries = flatten_to_entries(state)

    try:
        if template_path is not None:
            # Template mode: load user workbook and fill matched columns
            tpl = Path(template_path)
            try:
                wb = load_workbook(tpl)
            except Exception as load_exc:
                raise ExportError(
                    f"Failed to load template: {tpl}",
                    error_code="TEMPLATE_LOAD_FAILED",
                    details={"template_path": str(tpl)},
                ) from load_exc

            ws = wb.active
            if ws is None:
                raise ExportError(
                    "Template workbook has no active sheet",
                    error_code="TEMPLATE_LOAD_FAILED",
                    details={"template_path": str(tpl)},
                )
            _fill_template(ws, entries)
        else:
            # Default mode: create styled workbook from scratch
            wb = Workbook()
            ws = wb.active
            if ws is None:  # pragma: no cover — Workbook() always has active sheet
                raise ExportError(
                    "Failed to create workbook",
                    error_code="XLSX_EXPORT_FAILED",
                    details={"output_path": str(output)},
                )

            header_row = _write_metadata(ws, state, len(entries))
            _write_default_sheet(ws, entries, header_row)

        wb.save(output)
        return output

    except ExportError:
        raise
    except Exception as exc:
        raise ExportError(
            f"Failed to export xlsx: {exc}",
            error_code="XLSX_EXPORT_FAILED",
            details={
                "output_path": str(output),
                "rule_count": len(entries),
            },
        ) from exc


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

# Column definitions for PDF table: (field_name, display_header, relative_width)
_PDF_COLUMNS: list[tuple[str, str, int]] = [
    ("source", "Source", 16),
    ("destination", "Destination", 16),
    ("port", "Port", 7),
    ("protocol", "Protocol", 9),
    ("direction", "Direction", 11),
    ("action", "Action", 10),
    ("justification", "Justification", 26),
    ("risk", "Risk", 9),
    ("proposal_id", "Proposal ID", 13),
    ("approval_required", "Approval", 9),
]


def export_pdf(
    state: ExcelPipelineState,
    output_path: str | Path,
) -> Path:
    """Export pipeline state as a PDF change request document.

    Produces a styled PDF with a metadata header section and a table of
    decided rule proposals. Uses built-in Helvetica font only — no TTF
    embedding required.

    Args:
        state: ExcelPipelineState dict from pipeline execution.
        output_path: Destination path for the generated PDF file.

    Returns:
        Path to the written PDF file.

    Raises:
        ExportError: If PDF creation or writing fails.
    """
    output = Path(output_path)
    entries = flatten_to_entries(state)

    try:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        # --- Metadata header ---
        pdf.set_font("Helvetica", "B", 14)
        pdf.cell(text="Firewall Change Request", new_x="LMARGIN", new_y="NEXT")
        pdf.ln(4)

        pdf.set_font("Helvetica", "", 10)
        run_id = state.get("run_id", "")
        started_at = state.get("started_at", "")
        generated = started_at if started_at else datetime.now(timezone.utc).isoformat()

        pdf.cell(text=f"Generated: {generated}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(text=f"Run ID: {run_id}", new_x="LMARGIN", new_y="NEXT")
        pdf.cell(
            text="Source Type: Excel Traffic Analysis",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        pdf.cell(
            text=f"Total Rules: {len(entries)}",
            new_x="LMARGIN",
            new_y="NEXT",
        )
        pdf.ln(4)

        # Separator line
        page_w = pdf.w - pdf.l_margin - pdf.r_margin
        y = pdf.get_y()
        pdf.line(pdf.l_margin, y, pdf.l_margin + page_w, y)
        pdf.ln(4)

        # --- Table or empty message ---
        if not entries:
            pdf.set_font("Helvetica", "I", 10)
            pdf.cell(text="No rules proposed.", new_x="LMARGIN", new_y="NEXT")
        else:
            pdf.set_font("Helvetica", "", 7)

            headings_style = FontFace(
                emphasis="BOLD",
                color=255,
                fill_color=(68, 114, 196),
            )

            # Calculate column widths proportional to page width
            total_weight = sum(w for _, _, w in _PDF_COLUMNS)
            col_widths = tuple(
                round(page_w * w / total_weight, 1) for _, _, w in _PDF_COLUMNS
            )

            with pdf.table(
                col_widths=col_widths,
                borders_layout="SINGLE_TOP_LINE",
                headings_style=headings_style,
                line_height=pdf.font_size * 2,
            ) as table:
                # Header row
                header_row = table.row()
                for _, display_name, _ in _PDF_COLUMNS:
                    header_row.cell(display_name)

                # Data rows
                for entry in entries:
                    entry_dict = entry.model_dump()
                    data_row = table.row()
                    for field, _, _ in _PDF_COLUMNS:
                        value = entry_dict[field]
                        if isinstance(value, bool):
                            value = "Yes" if value else "No"
                        data_row.cell(str(value))

        pdf.output(str(output))
        return output

    except ExportError:
        raise
    except Exception as exc:
        raise ExportError(
            f"Failed to export PDF: {exc}",
            error_code="PDF_EXPORT_FAILED",
            details={
                "output_path": str(output),
                "rule_count": len(entries),
            },
        ) from exc
