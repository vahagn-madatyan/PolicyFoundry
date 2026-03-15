"""Excel traffic file ingestion.

Parses Excel exports of firewall traffic logs using openpyxl in read_only mode.
Auto-detects column layout via synonym matching (column_detect) or accepts an
explicit ColumnMapping override. Bad rows are skipped with warnings — never
raised as exceptions.
"""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl
from pydantic import ValidationError

from policyfoundry.exceptions import ExcelParseError
from policyfoundry.ingestion.column_detect import detect_columns
from policyfoundry.ingestion.excel_schema import (
    ColumnMapping,
    ExcelIngestionResult,
    ExcelTrafficRecord,
)

logger = logging.getLogger(__name__)


def _coerce_port(value: object) -> int | None:
    """Coerce a cell value to an integer port number.

    Handles int, float (Excel stores numbers as float), and string representations.
    Returns None if coercion fails.
    """
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value)
    if isinstance(value, str):
        stripped = value.strip()
        if stripped.isdigit():
            return int(stripped)
        # Try float-then-int for "443.0" style strings
        try:
            return int(float(stripped))
        except (ValueError, OverflowError):
            return None
    return None


def _cell_to_str(value: object) -> str | None:
    """Convert a cell value to a stripped string. Returns None for None/empty."""
    if value is None:
        return None
    s = str(value).strip()
    return s if s else None


def ingest_excel_file(
    path: str | Path,
    column_mapping: ColumnMapping | None = None,
    sheet_name: str | None = None,
    header_row: int = 1,
) -> ExcelIngestionResult:
    """Parse an Excel traffic export into ExcelTrafficRecords.

    Opens the workbook in read_only mode with explicit close via try/finally.
    Auto-detects columns from the header row unless a ColumnMapping is provided.
    Validation failures on individual rows are logged and counted as skipped —
    they never bubble up as exceptions.

    Args:
        path: Path to the Excel file.
        column_mapping: Pre-built column mapping. If None, auto-detect from headers.
        sheet_name: Sheet to read. Defaults to the first sheet.
        header_row: 1-based row number containing headers (default: 1).

    Returns:
        ExcelIngestionResult with records, stats, and any warnings.

    Raises:
        ExcelParseError: If the file is missing, unreadable, the sheet is not
            found, or column auto-detection fails.
    """
    file_path = Path(path)
    if not file_path.exists():
        raise ExcelParseError(
            f"Excel file not found: {file_path}",
            error_code="FILE_NOT_FOUND",
            details={"path": str(file_path)},
        )

    wb = None
    try:
        wb = openpyxl.load_workbook(
            file_path, read_only=True, data_only=True
        )

        # Select sheet
        if sheet_name is not None:
            if sheet_name not in wb.sheetnames:
                raise ExcelParseError(
                    f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}",
                    error_code="SHEET_NOT_FOUND",
                    details={
                        "requested_sheet": sheet_name,
                        "available_sheets": wb.sheetnames,
                    },
                )
            ws = wb[sheet_name]
        else:
            ws = wb[wb.sheetnames[0]]
            sheet_name = wb.sheetnames[0]

        # Read header row
        header_cells: list[str] = []
        for row in ws.iter_rows(
            min_row=header_row, max_row=header_row, values_only=True
        ):
            header_cells = [str(c).strip() if c is not None else "" for c in row]
            break

        if not header_cells or all(h == "" for h in header_cells):
            raise ExcelParseError(
                f"Header row {header_row} is empty or unreadable.",
                error_code="EMPTY_HEADER_ROW",
                details={"header_row": header_row},
            )

        # Detect or use provided column mapping
        if column_mapping is None:
            column_mapping = detect_columns(header_cells)
            logger.info("Auto-detected column mapping: %s", column_mapping)

        # Build field→index lookup for efficient row parsing
        field_indices: dict[str, int] = column_mapping.model_dump()

        # Parse data rows
        records: list[ExcelTrafficRecord] = []
        warnings: list[str] = []
        total_rows = 0
        skipped_rows = 0

        for row in ws.iter_rows(
            min_row=header_row + 1, values_only=True
        ):
            total_rows += 1

            # Skip completely empty rows
            if row is None or all(c is None for c in row):
                skipped_rows += 1
                warnings.append(f"Row {header_row + total_rows}: empty row skipped")
                continue

            try:
                # Extract values by column index
                row_len = len(row)
                raw: dict[str, object] = {}
                missing_cells: list[str] = []

                for field_name, col_idx in field_indices.items():
                    if col_idx >= row_len:
                        missing_cells.append(field_name)
                        continue
                    raw[field_name] = row[col_idx]

                if missing_cells:
                    skipped_rows += 1
                    warnings.append(
                        f"Row {header_row + total_rows}: missing cells for "
                        f"{', '.join(missing_cells)}"
                    )
                    continue

                # Coerce ports
                port1_val = _coerce_port(raw.get("port1"))
                port2_val = _coerce_port(raw.get("port2"))
                if port1_val is None or port2_val is None:
                    skipped_rows += 1
                    warnings.append(
                        f"Row {header_row + total_rows}: non-numeric port value "
                        f"(port1={raw.get('port1')!r}, port2={raw.get('port2')!r})"
                    )
                    continue

                # Convert string fields — skip if any required string is None/empty
                str_fields = {}
                has_none = False
                for f in (
                    "protocol",
                    "ip1",
                    "interface1",
                    "hostname1",
                    "ip2",
                    "interface2",
                    "hostname2",
                    "flag",
                ):
                    val = _cell_to_str(raw.get(f))
                    if val is None:
                        has_none = True
                        break
                    str_fields[f] = val

                if has_none:
                    skipped_rows += 1
                    warnings.append(
                        f"Row {header_row + total_rows}: None/empty cell in required field"
                    )
                    continue

                record = ExcelTrafficRecord(
                    port1=port1_val,
                    port2=port2_val,
                    **str_fields,
                )
                records.append(record)

            except ValidationError as exc:
                skipped_rows += 1
                logger.debug(
                    "Row %d validation failed: %s",
                    header_row + total_rows,
                    exc,
                )
                warnings.append(
                    f"Row {header_row + total_rows}: validation error — {exc.error_count()} issue(s)"
                )
            except Exception as exc:
                skipped_rows += 1
                logger.debug(
                    "Row %d unexpected error: %s",
                    header_row + total_rows,
                    exc,
                )
                warnings.append(
                    f"Row {header_row + total_rows}: unexpected error — {exc}"
                )

        return ExcelIngestionResult(
            records=records,
            column_mapping=column_mapping,
            total_rows=total_rows,
            parsed_rows=len(records),
            skipped_rows=skipped_rows,
            warnings=warnings,
            source_file=str(file_path),
        )

    except ExcelParseError:
        raise
    except Exception as exc:
        raise ExcelParseError(
            f"Failed to read Excel file: {exc}",
            error_code="FILE_READ_ERROR",
            details={"path": str(file_path), "original_error": str(exc)},
        ) from exc
    finally:
        if wb is not None:
            wb.close()
