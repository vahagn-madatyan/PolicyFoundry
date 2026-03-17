"""Tests for Excel file ingestion (ingest_excel_file).

Covers: successful parsing, column auto-detection integration, config override
column_mapping, missing file error, bad-row skipping, sheet_name selection,
and integration with the real 83K-row sample file.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from policyfoundry.exceptions import ExcelParseError
from policyfoundry.ingestion.excel import ingest_excel_file
from policyfoundry.ingestion.excel_schema import ColumnMapping

# ---------------------------------------------------------------------------
# Sample headers matching the real file's column order
# ---------------------------------------------------------------------------
SAMPLE_HEADERS = [
    "Protocol",
    "Interface1",
    "HostName1",
    "IP1",
    "Port1",
    "Interface2",
    "HostName2",
    "IP2",
    "Port2",
    "Flag",
]


def _make_row(
    protocol: str = "TCP",
    interface1: str = "inet",
    hostname1: str = "host1",
    ip1: str = "10.0.0.1",
    port1: int = 443,
    interface2: str = "zoneA",
    hostname2: str = "10.1.2.3",
    ip2: str = "10.1.2.3",
    port2: int = 53962,
    flag: str = "UIO",
) -> tuple:
    """Build a row tuple matching SAMPLE_HEADERS column order."""
    return (protocol, interface1, hostname1, ip1, port1, interface2, hostname2, ip2, port2, flag)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture()
def sample_excel(tmp_path: Path) -> Path:
    """Create a small Excel file with standard headers and 15 data rows.

    Includes edge cases:
    - Trailing whitespace on string cells
    - DNS annotation on hostname2
    - Port values as int and as string
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traffic"

    # Header row
    ws.append(SAMPLE_HEADERS)

    # 10 normal rows
    for i in range(10):
        ws.append(_make_row(
            ip1=f"10.0.0.{i + 1}",
            port1=443 + i,
            port2=50000 + i,
        ))

    # Row with trailing whitespace
    ws.append(_make_row(
        protocol="  UDP  ",
        ip1="  10.0.1.1  ",
        flag="  ACK  ",
    ))

    # Row with DNS annotation in hostname2
    ws.append(_make_row(
        hostname2="10.194.184.42 (no DNS resolution)",
    ))

    # Row with port as string
    ws.append(("TCP", "inet", "host1", "10.0.0.99", "8080", "zoneB", "host2", "10.0.0.100", "9090", "SYN"))

    # Row with port as float (Excel sometimes does this)
    ws.append(("TCP", "inet", "host1", "10.0.0.99", 443.0, "zoneB", "host2", "10.0.0.100", 80.0, "SYN"))

    # Row with None cell (will be skipped)
    ws.append(("TCP", "inet", None, "10.0.0.99", 443, "zoneB", "host2", "10.0.0.100", 80, "SYN"))

    file_path = tmp_path / "test_traffic.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture()
def multi_sheet_excel(tmp_path: Path) -> Path:
    """Excel file with two sheets — data on the second one."""
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Summary"
    ws1.append(["This sheet has no traffic data"])

    ws2 = wb.create_sheet("Firewall Logs")
    ws2.append(SAMPLE_HEADERS)
    for i in range(5):
        ws2.append(_make_row(ip1=f"192.168.1.{i + 1}"))

    file_path = tmp_path / "multi_sheet.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture()
def nonstandard_headers_excel(tmp_path: Path) -> Path:
    """Excel file with non-standard headers that won't auto-detect."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Proto", "SrcIntf", "SrcHost", "SrcAddr", "SrcP", "DstIntf", "DstHost", "DstAddr", "DstP", "Fl"])
    for i in range(3):
        ws.append(_make_row(ip1=f"172.16.0.{i + 1}"))

    file_path = tmp_path / "nonstandard.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


@pytest.fixture()
def bad_rows_excel(tmp_path: Path) -> Path:
    """Excel file with several unparseable rows interleaved with good ones."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SAMPLE_HEADERS)

    # Good row
    ws.append(_make_row())
    # Bad: port out of range (70000)
    ws.append(("TCP", "inet", "h1", "10.0.0.1", 70000, "zoneA", "h2", "10.0.0.2", 80, "SYN"))
    # Bad: non-numeric port
    ws.append(("TCP", "inet", "h1", "10.0.0.1", "abc", "zoneA", "h2", "10.0.0.2", 80, "SYN"))
    # Good row
    ws.append(_make_row(ip1="10.0.0.5"))
    # Bad: empty row
    ws.append((None, None, None, None, None, None, None, None, None, None))

    file_path = tmp_path / "bad_rows.xlsx"
    wb.save(file_path)
    wb.close()
    return file_path


# ---------------------------------------------------------------------------
# Tests: Successful parsing
# ---------------------------------------------------------------------------


class TestSuccessfulParsing:
    """Test successful ingestion with standard headers."""

    def test_parse_returns_correct_record_count(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        # 15 rows total: 10 normal + whitespace + dns + string port + float port + None cell
        # The None-cell row should be skipped
        assert result.parsed_rows == 14
        assert result.skipped_rows == 1
        assert result.total_rows == 15
        assert len(result.records) == 14

    def test_all_fields_populated(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        record = result.records[0]
        assert record.protocol == "TCP"
        assert record.ip1 == "10.0.0.1"
        assert record.port1 == 443
        assert record.interface1 == "inet"
        assert record.hostname1 == "host1"
        assert record.ip2 == "10.1.2.3"
        assert record.port2 == 50000
        assert record.interface2 == "zoneA"
        assert record.hostname2 == "10.1.2.3"
        assert record.flag == "UIO"

    def test_whitespace_stripped(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        # Row 11 (0-indexed: 10) has trailing whitespace
        ws_record = result.records[10]
        assert ws_record.protocol == "UDP"
        assert ws_record.ip1 == "10.0.1.1"
        assert ws_record.flag == "ACK"

    def test_dns_annotation_cleaned(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        # Row 12 has DNS annotation
        dns_record = result.records[11]
        assert dns_record.hostname2 == "10.194.184.42"

    def test_string_port_coerced(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        # Row 13 has string ports
        str_record = result.records[12]
        assert str_record.port1 == 8080
        assert str_record.port2 == 9090

    def test_float_port_coerced(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        # Row 14 has float ports
        float_record = result.records[13]
        assert float_record.port1 == 443
        assert float_record.port2 == 80

    def test_source_file_recorded(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        assert str(sample_excel) in result.source_file

    def test_column_mapping_populated(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        assert result.column_mapping is not None
        mapping = result.column_mapping.model_dump()
        assert set(mapping.keys()) == {
            "protocol",
            "ip1",
            "port1",
            "interface1",
            "hostname1",
            "ip2",
            "port2",
            "interface2",
            "hostname2",
            "flag",
        }


# ---------------------------------------------------------------------------
# Tests: Auto-detection integration
# ---------------------------------------------------------------------------


class TestAutoDetectIntegration:
    """Column auto-detection wired through ingest_excel_file."""

    def test_headers_detected_and_records_parsed(self, sample_excel: Path) -> None:
        result = ingest_excel_file(sample_excel)
        assert result.column_mapping is not None
        # Protocol is column 0 in our sample
        assert result.column_mapping.protocol == 0
        assert result.parsed_rows > 0

    def test_auto_detect_with_real_sample_headers(self, sample_excel: Path) -> None:
        """Verifies the real file's header order maps correctly."""
        result = ingest_excel_file(sample_excel)
        m = result.column_mapping
        assert m is not None
        # Matches SAMPLE_HEADERS order: Protocol, Interface1, HostName1, IP1, Port1, ...
        assert m.protocol == 0
        assert m.interface1 == 1
        assert m.hostname1 == 2
        assert m.ip1 == 3
        assert m.port1 == 4
        assert m.interface2 == 5
        assert m.hostname2 == 6
        assert m.ip2 == 7
        assert m.port2 == 8
        assert m.flag == 9


# ---------------------------------------------------------------------------
# Tests: Config override column_mapping
# ---------------------------------------------------------------------------


class TestColumnMappingOverride:
    """Explicit column_mapping bypasses auto-detection."""

    def test_override_with_standard_headers(self, sample_excel: Path) -> None:
        """Provide an explicit mapping matching the actual columns."""
        explicit = ColumnMapping(
            protocol=0,
            interface1=1,
            hostname1=2,
            ip1=3,
            port1=4,
            interface2=5,
            hostname2=6,
            ip2=7,
            port2=8,
            flag=9,
        )
        result = ingest_excel_file(sample_excel, column_mapping=explicit)
        assert result.parsed_rows == 14
        assert result.column_mapping == explicit

    def test_override_with_nonstandard_headers(
        self, nonstandard_headers_excel: Path
    ) -> None:
        """Non-standard headers that fail auto-detect work with explicit mapping."""
        explicit = ColumnMapping(
            protocol=0,
            interface1=1,
            hostname1=2,
            ip1=3,
            port1=4,
            interface2=5,
            hostname2=6,
            ip2=7,
            port2=8,
            flag=9,
        )
        result = ingest_excel_file(
            nonstandard_headers_excel, column_mapping=explicit
        )
        assert result.parsed_rows == 3


# ---------------------------------------------------------------------------
# Tests: Error handling
# ---------------------------------------------------------------------------


class TestErrorHandling:
    """Tests for error conditions."""

    def test_missing_file_raises_excel_parse_error(self, tmp_path: Path) -> None:
        with pytest.raises(ExcelParseError, match="not found"):
            ingest_excel_file(tmp_path / "nonexistent.xlsx")

    def test_missing_file_error_code(self, tmp_path: Path) -> None:
        with pytest.raises(ExcelParseError) as exc_info:
            ingest_excel_file(tmp_path / "nonexistent.xlsx")
        assert exc_info.value.error_code == "FILE_NOT_FOUND"

    def test_sheet_not_found_raises(self, sample_excel: Path) -> None:
        with pytest.raises(ExcelParseError, match="not found"):
            ingest_excel_file(sample_excel, sheet_name="DoesNotExist")

    def test_sheet_not_found_error_code(self, sample_excel: Path) -> None:
        with pytest.raises(ExcelParseError) as exc_info:
            ingest_excel_file(sample_excel, sheet_name="DoesNotExist")
        assert exc_info.value.error_code == "SHEET_NOT_FOUND"


# ---------------------------------------------------------------------------
# Tests: Bad row handling
# ---------------------------------------------------------------------------


class TestBadRowHandling:
    """Unparseable rows are skipped with warnings, not exceptions."""

    def test_bad_rows_skipped_with_warnings(self, bad_rows_excel: Path) -> None:
        result = ingest_excel_file(bad_rows_excel)
        assert result.parsed_rows == 2  # only the 2 good rows
        assert result.skipped_rows == 3  # port OOB, non-numeric port, empty row
        assert len(result.warnings) == 3

    def test_good_rows_still_parsed(self, bad_rows_excel: Path) -> None:
        result = ingest_excel_file(bad_rows_excel)
        ips = [r.ip1 for r in result.records]
        assert "10.0.0.1" in ips  # first good row
        assert "10.0.0.5" in ips  # second good row

    def test_total_rows_correct(self, bad_rows_excel: Path) -> None:
        result = ingest_excel_file(bad_rows_excel)
        assert result.total_rows == 5  # all rows counted


# ---------------------------------------------------------------------------
# Tests: Sheet selection
# ---------------------------------------------------------------------------


class TestSheetSelection:
    """Sheet name selection works correctly."""

    def test_explicit_sheet_name(self, multi_sheet_excel: Path) -> None:
        result = ingest_excel_file(multi_sheet_excel, sheet_name="Firewall Logs")
        assert result.parsed_rows == 5

    def test_default_uses_first_sheet(self, multi_sheet_excel: Path) -> None:
        """Without sheet_name, reads the first sheet (Summary) which has no traffic data."""
        # First sheet has non-standard data — should fail detection or have 0 records
        # depending on what the first sheet looks like
        with pytest.raises(ExcelParseError):
            ingest_excel_file(multi_sheet_excel)


# ---------------------------------------------------------------------------
# Integration: Real 83K-row sample file
# ---------------------------------------------------------------------------

REAL_SAMPLE = Path("examples/input/sample-traffic.xlsx")


@pytest.mark.skipif(
    not REAL_SAMPLE.exists(),
    reason="Real sample file not available",
)
class TestRealSampleIntegration:
    """Integration tests against the actual 83K-row sample file."""

    def test_parses_all_rows(self) -> None:
        result = ingest_excel_file(REAL_SAMPLE)
        assert result.parsed_rows == 83_633
        assert result.skipped_rows == 0

    def test_all_columns_detected(self) -> None:
        result = ingest_excel_file(REAL_SAMPLE)
        assert result.column_mapping is not None
        mapping = result.column_mapping.model_dump()
        assert len(mapping) == 10
        assert all(isinstance(v, int) for v in mapping.values())

    def test_first_record_fields(self) -> None:
        result = ingest_excel_file(REAL_SAMPLE)
        first = result.records[0]
        # Spot-check known values from row 2 of the real file
        assert first.protocol.strip() == "TCP"
        assert first.interface1 == "inet"
        assert first.port1 == 443
