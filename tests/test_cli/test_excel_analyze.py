"""CLI integration tests for `policyfoundry analyze --source excel`.

Tests cover the full Excel workflow:
- Rich formatted output with pipeline results
- JSON formatted output with all stage data
- --export xlsx produces a change request Excel file
- --export pdf produces a change request PDF file
- --export xlsx,pdf produces both
- --template custom.xlsx fills a user-provided template
- Error handling (missing --file, empty file, pipeline errors)
- Token cost display in output footer

All tests mock the LLM boundary — no real API calls.
"""

from __future__ import annotations

import json
from pathlib import Path
from typing import Any
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from typer.testing import CliRunner

from policyfoundry.adapters.schema import (
    Direction,
    NetworkEndpoint,
    PortRange,
    RiskLevel,
    RuleAction,
    UniversalRule,
)
from policyfoundry.exceptions import PipelineError, PolicyFoundryError
from policyfoundry.main import app
from policyfoundry.output.models import TokenUsage
from policyfoundry.pipeline.schema import PolicyProposal, RuleDecision


@pytest.fixture
def cli_runner() -> CliRunner:
    """Typer CLI test runner."""
    return CliRunner()


@pytest.fixture
def sample_excel_pipeline_state() -> dict[str, Any]:
    """Complete ExcelPipelineState dict with all stages populated.

    Matches the shape returned by run_excel_pipeline + token_usage attachment.
    """
    return {
        "run_id": "run-excel-cli-001",
        "started_at": "2026-03-15T09:00:00+00:00",
        "current_stage": "decide",
        "aggregated_flows": [
            {
                "src_ip": "10.0.1.50",
                "dst_ip": "10.0.2.10",
                "service_port": 443,
                "protocol": "TCP",
                "direction": "OUTBOUND",
                "flow_count": 800,
                "src_interface": "eth0",
                "dst_interface": "eth1",
                "sample_src_ports": [49152],
            },
            {
                "src_ip": "192.168.1.5",
                "dst_ip": "10.0.1.50",
                "service_port": 22,
                "protocol": "TCP",
                "direction": "INBOUND",
                "flow_count": 120,
                "src_interface": "eth1",
                "dst_interface": "eth0",
                "sample_src_ports": [50001],
            },
        ],
        "subnet_groups": [
            {
                "cidr": "10.0.1.0/24",
                "member_ips": ["10.0.1.50"],
                "member_count": 1,
                "shared_patterns": [{"protocol": "TCP", "port": 443}],
            },
        ],
        "analysis": {
            "summary": "Mixed traffic with HTTPS dominance.",
            "total_flows": 920,
            "unique_sources": 2,
            "unique_destinations": 2,
            "top_talkers": [
                {"ip": "10.0.1.50", "flow_count": 800, "protocol": "TCP"},
            ],
            "port_distribution": [
                {"port": 443, "protocol": "TCP", "percentage": 86.9},
                {"port": 22, "protocol": "TCP", "percentage": 13.1},
            ],
            "anomalies": [],
            "bandwidth_outliers": [],
        },
        "assessment": {
            "overall_risk": "MEDIUM",
            "risk_scores": [
                {
                    "category": "inferred_gaps",
                    "score": 0.5,
                    "description": "SSH access without explicit rule",
                },
            ],
            "rule_gaps": [
                {
                    "gap_type": "inferred_missing",
                    "description": "SSH from 192.168.1.5 likely needs rule",
                    "severity": "MEDIUM",
                },
            ],
            "compliance_findings": ["SSH access not restricted"],
        },
        "proposals": [
            {
                "proposal_id": "PROP-001",
                "rule": {
                    "name": "allow-https-outbound",
                    "description": "Allow HTTPS to API server",
                    "action": "ALLOW",
                    "direction": "OUTBOUND",
                    "protocol": "TCP",
                    "source": [{"cidr": "10.0.1.0/24"}],
                    "destination": [{"cidr": "10.0.2.10/32"}],
                    "port_range": {"from_port": 443, "to_port": 443},
                },
                "justification": "Repeated HTTPS traffic from subnet",
                "risk_level": "LOW",
                "confidence": 0.9,
                "impact_analysis": "Allows HTTPS from app subnet",
            },
        ],
        "decisions": [
            {
                "decision_id": "DEC-001",
                "proposal_id": "PROP-001",
                "action": "CREATE",
                "risk_level": "LOW",
                "reason": "Low-risk HTTPS rule approved",
                "approval_required": False,
            },
        ],
        "token_usage": {
            "prompt_tokens": 3000,
            "completion_tokens": 1200,
            "total_tokens": 4200,
            "total_cost": 0.0030,
            "per_stage": [
                {
                    "stage": "analyze",
                    "prompt_tokens": 1000,
                    "completion_tokens": 400,
                    "total_tokens": 1400,
                    "cost": 0.0010,
                },
            ],
        },
    }


@pytest.fixture
def mock_excel_file(tmp_path: Path) -> Path:
    """Create a minimal Excel file for CLI testing.

    Uses column headers that match the auto-detect synonyms in column_detect.py.
    """
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([
        "Protocol", "IP1", "Port1", "IP2", "Port2",
        "Interface1", "Interface2", "Hostname1", "Hostname2", "Flag",
    ])
    ws.append([
        "TCP", "10.0.1.50", "49152", "10.0.2.10", "443",
        "eth0", "eth1", "host-a", "host-b", "UI",
    ])
    ws.append([
        "TCP", "192.168.1.5", "50001", "10.0.1.50", "22",
        "eth1", "eth0", "host-c", "host-a", "UIO",
    ])

    excel_path = tmp_path / "traffic.xlsx"
    wb.save(excel_path)
    return excel_path


@pytest.fixture
def mock_ingestion_result():
    """Create a mock ExcelIngestionResult with minimal valid records."""
    from policyfoundry.ingestion.excel_schema import (
        ColumnMapping,
        ExcelIngestionResult,
        ExcelTrafficRecord,
    )

    records = [
        ExcelTrafficRecord(
            protocol="TCP",
            ip1="10.0.1.50",
            port1=49152,
            ip2="10.0.2.10",
            port2=443,
            interface1="eth0",
            interface2="eth1",
            hostname1="host-a",
            hostname2="host-b",
            flag="UI",
        ),
        ExcelTrafficRecord(
            protocol="TCP",
            ip1="192.168.1.5",
            port1=50001,
            ip2="10.0.1.50",
            port2=22,
            interface1="eth1",
            interface2="eth0",
            hostname1="host-c",
            hostname2="host-a",
            flag="UIO",
        ),
    ]

    return ExcelIngestionResult(
        records=records,
        column_mapping=ColumnMapping(
            protocol=0, ip1=1, port1=2, ip2=3, port2=4,
            interface1=5, interface2=6, hostname1=7, hostname2=8, flag=9,
        ),
        total_rows=2,
        parsed_rows=2,
        skipped_rows=0,
        warnings=[],
        source_file="traffic.xlsx",
    )


def _mock_config():
    """Create a default PolicyFoundryConfig for Excel tests."""
    from policyfoundry.config.models import PolicyFoundryConfig

    return PolicyFoundryConfig()


def _mock_llm_client(pipeline_state: dict):
    """Create mock LLM client that returns pipeline_state from run_excel_pipeline."""
    mock_usage = TokenUsage(
        prompt_tokens=3000,
        completion_tokens=1200,
        total_tokens=4200,
        total_cost=0.0030,
    )

    mock_llm = MagicMock()
    mock_llm.get_usage.return_value = mock_usage
    return mock_llm


def _excel_patches(pipeline_state, mock_ingestion_result_val):
    """Return standard patches for Excel CLI tests.

    Mocks: load_config, create_llm_client, ingest_excel_file, run_excel_pipeline.
    """
    state = {k: v for k, v in pipeline_state.items() if k != "token_usage"}
    mock_llm = _mock_llm_client(pipeline_state)

    return (
        patch("policyfoundry.main.load_config", return_value=_mock_config()),
        patch("policyfoundry.main.create_llm_client", new_callable=AsyncMock, return_value=mock_llm),
        patch("policyfoundry.ingestion.excel.ingest_excel_file", return_value=mock_ingestion_result_val),
        patch("policyfoundry.main.run_excel_pipeline", new_callable=AsyncMock, return_value=state),
    )


# ---------------------------------------------------------------------------
# Rich output tests
# ---------------------------------------------------------------------------


class TestExcelAnalyzeRichOutput:
    """Tests for `policyfoundry analyze --source excel` with Rich output."""

    def test_excel_rich_output_exits_zero(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Excel analyze exits 0 with Rich output."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"

    def test_excel_rich_output_contains_pipeline_summary(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Rich output includes Excel Pipeline Summary panel."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
            ])
            assert "Excel Pipeline Summary" in result.output

    def test_excel_rich_output_contains_decisions(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Rich output includes Decisions section."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
            ])
            assert "Decisions" in result.output or "CREATE" in result.output

    def test_excel_rich_output_contains_token_usage(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Rich output includes Token Usage footer."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
            ])
            assert "Token Usage" in result.output


# ---------------------------------------------------------------------------
# JSON output tests
# ---------------------------------------------------------------------------


class TestExcelAnalyzeJsonOutput:
    """Tests for `policyfoundry analyze --source excel --format json`."""

    def test_excel_json_output_exits_zero(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Excel analyze with --format json exits 0."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--format", "json",
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"

    def test_excel_json_output_is_valid_json(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """JSON output parses as valid JSON with expected keys."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--format", "json",
            ])
            parsed = json.loads(result.output)
            assert isinstance(parsed, dict)
            assert "run_id" in parsed
            assert "aggregated_flows" in parsed
            assert "proposals" in parsed
            assert "decisions" in parsed

    def test_excel_json_output_contains_token_usage(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """JSON output includes token_usage field."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--format", "json",
            ])
            parsed = json.loads(result.output)
            assert "token_usage" in parsed
            assert parsed["token_usage"]["total_tokens"] == 4200


# ---------------------------------------------------------------------------
# Export tests
# ---------------------------------------------------------------------------


class TestExcelAnalyzeExport:
    """Tests for --export xlsx/pdf options."""

    def test_export_xlsx_creates_file(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """--export xlsx creates a .xlsx change request file."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "xlsx",
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"
            assert "Excel change request exported" in result.output

            expected_path = mock_excel_file.parent / "traffic_change_request.xlsx"
            assert expected_path.exists(), f"Expected {expected_path} to exist"

    def test_export_pdf_creates_file(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """--export pdf creates a .pdf change request file."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "pdf",
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"
            assert "PDF change request exported" in result.output

            expected_path = mock_excel_file.parent / "traffic_change_request.pdf"
            assert expected_path.exists(), f"Expected {expected_path} to exist"

    def test_export_both_xlsx_and_pdf(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """--export xlsx,pdf creates both files."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "xlsx,pdf",
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"
            assert "Excel change request exported" in result.output
            assert "PDF change request exported" in result.output

            xlsx_path = mock_excel_file.parent / "traffic_change_request.xlsx"
            pdf_path = mock_excel_file.parent / "traffic_change_request.pdf"
            assert xlsx_path.exists()
            assert pdf_path.exists()

    def test_export_xlsx_with_template(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result, tmp_path,
    ):
        """--export xlsx --template custom.xlsx fills the template."""
        from openpyxl import Workbook

        tpl_wb = Workbook()
        tpl_ws = tpl_wb.active
        tpl_ws.append(["Source", "Destination", "Port", "Protocol", "Direction",
                        "Action", "Justification", "Risk"])
        tpl_path = tmp_path / "custom_template.xlsx"
        tpl_wb.save(tpl_path)

        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "xlsx", "--template", str(tpl_path),
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"
            assert "Excel change request exported" in result.output

    def test_export_xlsx_contains_rule_data(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Exported xlsx contains actual rule data from pipeline state."""
        from openpyxl import load_workbook

        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "xlsx",
            ])
            assert result.exit_code == 0

        xlsx_path = mock_excel_file.parent / "traffic_change_request.xlsx"
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Metadata in row 4 shows rule count
        total_rules_cell = ws.cell(row=4, column=2).value
        assert total_rules_cell == 1  # One non-SKIP decision

        # Data starts after header row 6 → row 7
        source_cell = ws.cell(row=7, column=1).value
        assert source_cell is not None  # Source should be populated

    def test_export_pdf_has_valid_magic_bytes(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Exported PDF starts with %PDF magic bytes."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "pdf",
            ])
            assert result.exit_code == 0

        pdf_path = mock_excel_file.parent / "traffic_change_request.pdf"
        with open(pdf_path, "rb") as f:
            magic = f.read(4)
        assert magic == b"%PDF"


# ---------------------------------------------------------------------------
# Error handling tests
# ---------------------------------------------------------------------------


class TestExcelAnalyzeErrorHandling:
    """Tests for Excel analyze error paths."""

    def test_missing_file_option_shows_error(self, cli_runner):
        """--source excel without --file shows actionable error."""
        with patch("policyfoundry.main.load_config", return_value=_mock_config()):
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel",
            ])
            assert result.exit_code == 1
            assert "MISSING_FILE_OPTION" in result.output

    def test_template_without_export_shows_error(self, cli_runner, mock_excel_file, tmp_path):
        """--template without --export xlsx shows actionable error."""
        tpl_path = tmp_path / "template.xlsx"
        tpl_path.touch()

        with patch("policyfoundry.main.load_config", return_value=_mock_config()):
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--template", str(tpl_path),
            ])
            assert result.exit_code == 1
            assert "TEMPLATE_WITHOUT_EXPORT" in result.output

    def test_pipeline_error_shows_actionable_message(
        self, cli_runner, mock_excel_file, mock_ingestion_result,
    ):
        """PipelineError produces exit 1 with structured error panel."""
        mock_llm = MagicMock()
        mock_llm.get_usage.return_value = TokenUsage()

        with (
            patch("policyfoundry.main.load_config", return_value=_mock_config()),
            patch("policyfoundry.main.create_llm_client", new_callable=AsyncMock, return_value=mock_llm),
            patch("policyfoundry.ingestion.excel.ingest_excel_file", return_value=mock_ingestion_result),
            patch(
                "policyfoundry.main.run_excel_pipeline",
                new_callable=AsyncMock,
                side_effect=PipelineError(
                    "Excel pipeline failed at stage: analyze",
                    error_code="PIPELINE_STAGE_FAILED",
                    details={"stage": "analyze"},
                ),
            ),
        ):
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
            ])
            assert result.exit_code == 1
            assert "PipelineError" in result.output
            assert "PIPELINE_STAGE_FAILED" in result.output

    def test_nonexistent_file_shows_error(self, cli_runner, tmp_path):
        """Non-existent Excel file shows error."""
        fake_path = tmp_path / "nonexistent.xlsx"

        with patch("policyfoundry.main.load_config", return_value=_mock_config()):
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(fake_path),
            ])
            assert result.exit_code == 1


# ---------------------------------------------------------------------------
# End-to-end composition test
# ---------------------------------------------------------------------------


class TestExcelAnalyzeEndToEnd:
    """Tests that verify all layers compose correctly."""

    def test_full_pipeline_rich_then_export(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """Full workflow: Rich output + xlsx + pdf export in one invocation."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--export", "xlsx,pdf",
            ])
            assert result.exit_code == 0, f"Exit {result.exit_code}: {result.output}"

            # Rich output was displayed
            assert "Excel Pipeline Summary" in result.output
            assert "Token Usage" in result.output

            # Both exports were created
            assert "Excel change request exported" in result.output
            assert "PDF change request exported" in result.output

            xlsx_path = mock_excel_file.parent / "traffic_change_request.xlsx"
            pdf_path = mock_excel_file.parent / "traffic_change_request.pdf"
            assert xlsx_path.exists()
            assert pdf_path.exists()

    def test_json_output_with_export(
        self, cli_runner, mock_excel_file, sample_excel_pipeline_state, mock_ingestion_result,
    ):
        """JSON output format + export work together."""
        p1, p2, p3, p4 = _excel_patches(sample_excel_pipeline_state, mock_ingestion_result)
        with p1, p2, p3, p4:
            result = cli_runner.invoke(app, [
                "analyze", "--source", "excel", "--file", str(mock_excel_file),
                "--format", "json", "--export", "xlsx",
            ])
            assert result.exit_code == 0

            xlsx_path = mock_excel_file.parent / "traffic_change_request.xlsx"
            assert xlsx_path.exists()
