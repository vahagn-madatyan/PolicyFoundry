"""Tests for xlsx change request export."""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from policyfoundry.exceptions import ExportError
from policyfoundry.export.change_request import export_xlsx
from policyfoundry.pipeline.excel_state import ExcelPipelineState


class TestDefaultExport:
    """Tests for default (no template) xlsx export."""

    def test_creates_valid_xlsx(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Default export creates a readable xlsx file."""
        out = tmp_path / "output.xlsx"
        result = export_xlsx(sample_excel_state, out)

        assert result == out
        assert out.exists()

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

    def test_metadata_rows(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Metadata rows contain expected labels and values."""
        out = tmp_path / "output.xlsx"
        export_xlsx(sample_excel_state, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

        assert ws.cell(row=1, column=1).value == "Generated"
        assert ws.cell(row=1, column=2).value == "2026-03-15T09:00:00+00:00"
        assert ws.cell(row=2, column=1).value == "Run ID"
        assert ws.cell(row=2, column=2).value == "run-export-test-001"
        assert ws.cell(row=3, column=1).value == "Source Type"
        assert ws.cell(row=3, column=2).value == "Excel Pipeline"
        assert ws.cell(row=4, column=1).value == "Total Rules"
        assert ws.cell(row=4, column=2).value == 2  # 2 non-SKIP entries

    def test_header_row_content(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Header row (row 6) contains expected column names."""
        out = tmp_path / "output.xlsx"
        export_xlsx(sample_excel_state, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

        header_row = 6
        expected_headers = [
            "Source", "Destination", "Port", "Protocol", "Direction",
            "Action", "Justification", "Risk", "Proposal ID", "Approval Required",
        ]
        actual_headers = [ws.cell(row=header_row, column=i).value for i in range(1, 11)]
        assert actual_headers == expected_headers

    def test_data_rows(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Data rows contain correct entry values."""
        out = tmp_path / "output.xlsx"
        export_xlsx(sample_excel_state, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

        data_start = 7  # row after header at 6

        # First data row — prop-001
        assert ws.cell(row=data_start, column=1).value == "10.0.1.0/24, 10.0.2.0/24"
        assert ws.cell(row=data_start, column=2).value == "10.0.3.10/32"
        assert ws.cell(row=data_start, column=3).value == "22"
        assert ws.cell(row=data_start, column=6).value == "CREATE"
        assert ws.cell(row=data_start, column=10).value == "Yes"

        # Second data row — prop-002
        assert ws.cell(row=data_start + 1, column=1).value == "any"
        assert ws.cell(row=data_start + 1, column=6).value == "UPDATE"
        assert ws.cell(row=data_start + 1, column=10).value == "No"

    def test_column_count(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Default export has exactly 10 columns."""
        out = tmp_path / "output.xlsx"
        export_xlsx(sample_excel_state, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

        header_row = 6
        # Count non-None cells in header row
        col_count = sum(
            1 for i in range(1, 20)
            if ws.cell(row=header_row, column=i).value is not None
        )
        assert col_count == 10


class TestTemplateExport:
    """Tests for custom template xlsx export."""

    def test_fills_template_columns(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Data fills into template columns matched by header name."""
        # Create a template with columns in a different order
        tpl = tmp_path / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1, value="Action")
        ws.cell(row=1, column=2, value="Source")
        ws.cell(row=1, column=3, value="Risk")
        ws.cell(row=1, column=4, value="Port")
        ws.cell(row=1, column=5, value="Unrelated Column")
        wb.save(tpl)

        out = tmp_path / "filled.xlsx"
        export_xlsx(sample_excel_state, out, template_path=tpl)

        wb2 = load_workbook(out)
        ws2 = wb2.active
        assert ws2 is not None

        # Row 2 should have first entry data in template column order
        assert ws2.cell(row=2, column=1).value == "CREATE"  # Action
        assert ws2.cell(row=2, column=2).value == "10.0.1.0/24, 10.0.2.0/24"  # Source
        assert ws2.cell(row=2, column=3).value == "MEDIUM"  # Risk
        assert ws2.cell(row=2, column=4).value == "22"  # Port
        assert ws2.cell(row=2, column=5).value is None  # Unrelated — untouched

        # Row 3 should have second entry
        assert ws2.cell(row=3, column=1).value == "UPDATE"
        assert ws2.cell(row=3, column=2).value == "any"

    def test_template_case_insensitive_headers(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Template header matching is case-insensitive."""
        tpl = tmp_path / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1, value="PROTOCOL")
        ws.cell(row=1, column=2, value="Direction")
        wb.save(tpl)

        out = tmp_path / "filled.xlsx"
        export_xlsx(sample_excel_state, out, template_path=tpl)

        wb2 = load_workbook(out)
        ws2 = wb2.active
        assert ws2 is not None

        assert ws2.cell(row=2, column=1).value == "TCP"
        assert ws2.cell(row=2, column=2).value == "INBOUND"


class TestEmptyProposals:
    """Tests for export with no proposals/decisions."""

    def test_default_empty(
        self,
        sample_excel_state_empty: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Empty state produces metadata-only xlsx (no data rows)."""
        out = tmp_path / "empty.xlsx"
        export_xlsx(sample_excel_state_empty, out)

        wb = load_workbook(out)
        ws = wb.active
        assert ws is not None

        # Metadata present
        assert ws.cell(row=1, column=1).value == "Generated"
        assert ws.cell(row=4, column=2).value == 0  # Total Rules = 0

        # Header present, but no data rows
        assert ws.cell(row=6, column=1).value == "Source"
        assert ws.cell(row=7, column=1).value is None  # No data row

    def test_template_empty(
        self,
        sample_excel_state_empty: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Empty state with template produces no data rows."""
        tpl = tmp_path / "template.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1, value="Source")
        ws.cell(row=1, column=2, value="Action")
        wb.save(tpl)

        out = tmp_path / "empty_filled.xlsx"
        export_xlsx(sample_excel_state_empty, out, template_path=tpl)

        wb2 = load_workbook(out)
        ws2 = wb2.active
        assert ws2 is not None

        # Headers intact, no data
        assert ws2.cell(row=1, column=1).value == "Source"
        assert ws2.cell(row=2, column=1).value is None


class TestExportErrors:
    """Tests for ExportError handling."""

    def test_invalid_output_path(
        self,
        sample_excel_state: ExcelPipelineState,
    ) -> None:
        """Writing to an invalid path raises ExportError."""
        with pytest.raises(ExportError) as exc_info:
            export_xlsx(sample_excel_state, "/nonexistent/dir/output.xlsx")

        assert exc_info.value.error_code == "XLSX_EXPORT_FAILED"

    def test_invalid_template_path(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Loading a nonexistent template raises ExportError with TEMPLATE_LOAD_FAILED."""
        out = tmp_path / "output.xlsx"
        with pytest.raises(ExportError) as exc_info:
            export_xlsx(
                sample_excel_state, out,
                template_path="/nonexistent/template.xlsx",
            )

        assert exc_info.value.error_code == "TEMPLATE_LOAD_FAILED"

    def test_template_no_matching_columns(
        self,
        sample_excel_state: ExcelPipelineState,
        tmp_path: Path,
    ) -> None:
        """Template with only unrecognized columns raises ExportError."""
        tpl = tmp_path / "bad_template.xlsx"
        wb = Workbook()
        ws = wb.active
        assert ws is not None
        ws.cell(row=1, column=1, value="FakeCol1")
        ws.cell(row=1, column=2, value="FakeCol2")
        ws.cell(row=1, column=3, value="NotARealColumn")
        wb.save(tpl)

        out = tmp_path / "output.xlsx"
        with pytest.raises(ExportError) as exc_info:
            export_xlsx(sample_excel_state, out, template_path=tpl)

        assert exc_info.value.error_code == "TEMPLATE_NO_MATCHING_COLUMNS"
